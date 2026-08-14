#!/usr/bin/env python3
"""
make_json.py -- reduced dashboard data generator.

Produces ONLY the two JSON files the view-only dashboard needs:
    reports.json  <-  Updates tab (A:E) of the accounting package
    flags.json    <-  "Deal Detail Calcs" + "_Master Inventory DIT" tabs

It reads ONE workbook (the AP Acct Pckg) and writes nothing back. The extraction
logic is copied verbatim from accounting_sync.py, so the output is byte-for-byte
identical to what the full sync produced -- this script just drops everything
that isn't needed to make these two files.

Why the --recalc toggle exists
------------------------------
flags.json reads FORMULA cells. openpyxl has no formula engine: with
data_only=True it returns the value that was *last cached in the file* by
whatever program calculated + saved it. Two cases:

  * The package is saved by Excel (or was already recalced) -> cached values are
    fresh -> the DEFAULT path (openpyxl only, no --recalc) is correct and fast.
  * You can't rely on a recent Excel save -> pass --recalc and LibreOffice will
    recompute a temp copy first, then openpyxl reads that. Heavier, but safe.

Because this script never writes the workbook, it doesn't wipe cached values the
way the full sync did -- so the openpyxl-only path is usually enough.

Usage
-----
    python make_json.py --workbook "/path/AP Acct Pckg .v6 - Live.xlsx"
    python make_json.py --reports-dir "/path/Reports"          # auto-discovers
    python make_json.py --workbook pkg.xlsx --out ./data --recalc

Env equivalents: WORKBOOK, REPORTS_DIR, MAKE_JSON_OUT, MAKE_JSON_RECALC=1
"""
import os, sys, glob, json, re, shutil, tempfile, subprocess, argparse
import datetime as dt

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Missing dependency: openpyxl  (pip install openpyxl)")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)   # repo root (this script lives in build/)

# ---------------------------------------------------------------- CLI / env ---
ap = argparse.ArgumentParser(
    description="Generate reports.json + flags.json from the accounting package.")
ap.add_argument("--workbook", default=os.environ.get("WORKBOOK"),
                help="Path to the AP Acct Pckg .xlsx to read.")
ap.add_argument("--reports-dir", default=os.environ.get("REPORTS_DIR"),
                help="Folder to auto-discover 'AP Acct Pckg*Live.xlsx' when "
                     "--workbook is omitted.")
ap.add_argument("--out", default=os.environ.get("MAKE_JSON_OUT",
                                                 os.path.join(ROOT, "data")),
                help="Output folder for the JSON (default: <repo>/data).")
_recalc_env = os.environ.get("MAKE_JSON_RECALC", "").lower() in ("1", "true", "yes", "on")
ap.add_argument("--recalc", dest="recalc", action="store_true", default=_recalc_env,
                help="Recompute formulas with LibreOffice before reading "
                     "(use if the workbook's cached values may be stale).")
ap.add_argument("--no-recalc", dest="recalc", action="store_false",
                help="Read cached values as-is (openpyxl only; the default).")
args = ap.parse_args()


def die(msg):
    sys.exit("ERROR: " + msg)


# ------------------------------------------------------- locate the workbook ---
def find_package(reports_dir):
    """Newest 'AP Acct Pckg*Live.xlsx' in reports_dir that isn't a backup/lock."""
    cands = [f for f in glob.glob(os.path.join(reports_dir, "AP Acct Pckg*Live.xlsx"))
             if "backup" not in f.lower() and not os.path.basename(f).startswith("~$")]
    return max(cands, key=os.path.getmtime) if cands else None


wb_path = args.workbook
if not wb_path and args.reports_dir:
    wb_path = find_package(args.reports_dir)
if not wb_path:
    die("no workbook given. Pass --workbook PATH, or --reports-dir DIR, "
        "or set the WORKBOOK env var.")
if not os.path.isfile(wb_path):
    die("workbook not found: %s" % wb_path)


# --------------------------------------------- optional LibreOffice recalc ---
# (verbatim from accounting_sync.py: _find_soffice / _recalc_copy)
def find_soffice():
    for cand in ("soffice", "libreoffice",
                 "/Applications/LibreOffice.app/Contents/MacOS/soffice",
                 "/usr/bin/soffice"):
        p = shutil.which(cand) if not os.path.isabs(cand) else \
            (cand if os.path.exists(cand) else None)
        if p:
            return p
    return None


def recalc_copy(src_path, soffice):
    outdir = tempfile.mkdtemp(prefix="makejson_recalc_")
    profile = "file://" + os.path.join(outdir, "lo_profile")
    try:
        subprocess.run([soffice, "-env:UserInstallation=" + profile, "--headless",
                        "--calc", "--convert-to", "xlsx", "--outdir", outdir, src_path],
                       timeout=240, stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL, check=True)
    except Exception:
        return None
    out = os.path.join(outdir,
                       os.path.splitext(os.path.basename(src_path))[0] + ".xlsx")
    return out if os.path.isfile(out) else None


read_path = wb_path
if args.recalc:
    soffice = find_soffice()
    if not soffice:
        print("WARNING: --recalc requested but LibreOffice not found; reading "
              "cached values as-is (flags.json may be blank if formulas were "
              "never computed).", file=sys.stderr)
    else:
        rc = recalc_copy(wb_path, soffice)
        if rc:
            read_path = rc
            print("recalc: LibreOffice recomputed a copy ->", rc)
        else:
            print("WARNING: LibreOffice recalc failed; reading cached values as-is.",
                  file=sys.stderr)

wb = load_workbook(read_path, data_only=True)
now = dt.datetime.now().isoformat(timespec="seconds")
os.makedirs(args.out, exist_ok=True)


# ----------------------------- formatting helpers (verbatim from the sync) ---
def _iso(v):
    return v.isoformat() if isinstance(v, (dt.datetime, dt.date)) else v


def fmt(v):
    if isinstance(v, (dt.datetime, dt.date)):
        try:
            return v.strftime("%-m/%-d/%y")
        except Exception:
            return v.strftime("%m/%d/%y")
    if isinstance(v, dt.time):
        return ""
    if isinstance(v, float):
        return (("%s" % format(v, ",.0f")) if v == int(v) else ("%s" % format(v, ",.2f")))
    return "" if v is None else str(v)


def split_title(full):
    full = str(full or "")
    name = full.split(":")[0].split(" (Amount")[0].strip()
    desc = full.split(":", 1)[1].strip() if ":" in full else ""
    return name, desc


# ----------------------------------------------- reports.json  (Updates A:E) ---
try:
    u = wb["UPDATES"] if "UPDATES" in wb.sheetnames else wb["Updates"]
    cols = [u.cell(1, c).value for c in range(1, 6)]
    rows = [[_iso(u.cell(r, c).value) for c in range(1, 6)]
            for r in range(2, u.max_row + 1)
            if any(u.cell(r, c).value not in (None, "") for c in range(1, 6))]
    json.dump({"generated": now, "columns": cols, "rows": rows},
              open(os.path.join(args.out, "reports.json"), "w"), indent=1)
    print("reports.json:", len(rows), "rows")
except Exception as e:
    print("reports.json skipped (%s)" % e, file=sys.stderr)


# ------------------------- flags.json  (Deal Detail Calcs + Master Inv DIT) ---
try:
    t = wb["Deal Detail Calcs"]
    secs = []
    for c0, ncol in ((1, 4), (6, 2), (9, 2), (12, 2), (15, 3), (19, 3), (23, 4)):
        title, desc = split_title(t.cell(1, c0).value)
        if not desc:
            desc = re.sub(r'^\s*\d+[a-z]?\.\s*', '', str(t.cell(2, c0).value or "").strip())
        hdrs = [str(t.cell(3, c0 + i).value or "") for i in range(ncol)]
        hdrs = ["Deal Age" if h == "Days Since Sale" else h for h in hdrs]
        rows = []
        for r in range(4, 500):
            if t.cell(r, c0).value in (None, ""):
                break
            rows.append([fmt(t.cell(r, c0 + i).value) for i in range(ncol)])
        if not rows or (len(rows) == 1 and str(rows[0][0]).lower() == "none found"):
            continue
        secs.append({"title": title, "desc": desc, "headers": hdrs, "rows": rows,
                     "source": "Deal Detail Calcs"})
    m = wb["_Master Inventory DIT"]
    for c in (1, 3, 5, 7, 9, 11, 13):
        title, desc = split_title(m.cell(26, c).value)
        field = str(m.cell(27, c).value or "").strip()
        if field and title.rstrip().lower().endswith("of"):
            title = (title.rstrip() + " " + field).strip()
            if not desc:
                desc = "Cars in inventory missing their %s in Dealr.Cloud - clean up at the source." % field
        data = [fmt(m.cell(r, c).value) for r in range(28, m.max_row + 1)
                if m.cell(r, c).value not in (None, "") and str(m.cell(r, c).value).lower() != "none"]
        if data:
            secs.append({"title": title, "desc": desc, "headers": ["Stock #"],
                         "rows": [[x] for x in data], "source": "Master Inventory DIT"})
    mt_h = [str(m.cell(27, c).value or "") for c in (15, 16, 17)]
    mt_r = [[fmt(m.cell(r, c).value) for c in (15, 16, 17)]
            for r in range(28, m.max_row + 1) if m.cell(r, 15).value not in (None, "")]
    if mt_r:
        nm, ds = split_title(m.cell(26, 15).value)
        secs.append({"title": nm, "desc": ds, "headers": mt_h, "rows": mt_r,
                     "source": "Master Inventory DIT"})
    for s in secs:
        tl = s["title"].lower()
        s["group"] = "Titles" if "title" in tl else \
            ("Deals" if s["source"] == "Deal Detail Calcs" else "Inventory")
    json.dump({"generated": now, "sections": secs},
              open(os.path.join(args.out, "flags.json"), "w"), indent=1)
    print("flags.json:", len(secs), "sections")
except Exception as e:
    print("flags.json skipped (%s)" % e, file=sys.stderr)
